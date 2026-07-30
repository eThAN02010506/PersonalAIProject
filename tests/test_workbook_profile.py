import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from qwopus_agent.analysis.excel_processing import read_spreadsheet
from qwopus_agent.analysis.workbook_profile import inspect_workbook


class WorkbookProfileTests(unittest.TestCase):
    def test_detects_single_numeric_column_header(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            pd.DataFrame({"score": [8, 9, 10]}).to_excel(path, index=False)

            result = read_spreadsheet(path)

        # 原因：单列样本是置信区间和单样本 t 检验的常见输入。
        # 作用：确认字段名不会被当作首条数据，统计 Skill 可以按真实列名取数。
        self.assertEqual(list(result.analysis_frames()["Sheet1"].columns), ["score"])
        self.assertEqual(result.analysis_frames()["Sheet1"]["score"].tolist(), [8, 9, 10])

    def test_equal_header_scores_prefer_the_earliest_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "metadata.xlsx"
            pd.DataFrame(
                [
                    ["Country Code", "Region", "Income Group"],
                    ["AAA", "Region A", "High income"],
                    ["BBB", "Region B", "Low income"],
                    ["CCC", "Region C", "Middle income"],
                ]
            ).to_excel(path, index=False, header=False)

            profile = inspect_workbook(path)

        # 原因：全字符串数据行可与字段行同分，后部行不能覆盖真正的首行表头。
        # 作用：确保元数据字段可被 Agent 识别并用于同层级筛选。
        self.assertEqual(profile.sheets[0].primary_region().header_rows, (1,))

    def test_detects_and_flattens_two_row_header(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "multi_header.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sales"
            sheet.merge_cells("A1:B1")
            sheet["A1"] = "Sales"
            sheet["C1"] = "Meta"
            sheet.append(["Region", "Revenue", "Owner"])
            sheet.append(["East", 10, "Alice"])
            sheet.append(["West", 20, "Bob"])
            workbook.save(path)

            result = read_spreadsheet(path)

        profile = result.profile
        self.assertIsNotNone(profile)
        assert profile is not None
        sheet_profile = profile.sheet("Sales")
        self.assertEqual(sheet_profile.kind, "table")
        self.assertEqual(sheet_profile.primary_region().header_rows, (1, 2))
        self.assertEqual(
            list(result.sheets["Sales"].columns),
            ["Sales | Region", "Sales | Revenue", "Meta | Owner"],
        )
        self.assertEqual(result.sheets["Sales"].iloc[0, 1], 10)

    def test_separates_multiple_vertical_tables(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "multiple_tables.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(["region", "revenue"])
            sheet.append(["East", 10])
            sheet.append(["West", 20])
            sheet.append([])
            sheet.append(["team", "tickets"])
            sheet.append(["Alpha", 4])
            sheet.append(["Beta", 7])
            workbook.save(path)

            result = read_spreadsheet(path)

        assert result.profile is not None
        sheet_profile = result.profile.sheet("Data")
        self.assertEqual(sheet_profile.kind, "multi_table")
        self.assertEqual(len(sheet_profile.table_regions), 2)
        self.assertEqual(
            set(result.region_sheets["Data"]),
            {"table_1", "table_2"},
        )
        self.assertEqual(
            list(result.region_sheets["Data"]["table_2"].columns),
            ["team", "tickets"],
        )
        self.assertEqual(
            set(result.analysis_frames()),
            {"Data", "Data::table_2"},
        )

    def test_form_sheet_uses_generic_columns_instead_of_false_header(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "form.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Profile"
            for row, (label, value) in enumerate(
                (
                    ("Name", "Ada"),
                    ("Role", "Engineer"),
                    ("Office", "London"),
                ),
                start=1,
            ):
                sheet.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=2,
                )
                sheet.cell(row, 1, label)
                sheet.cell(row, 3, value)
            workbook.save(path)

            result = read_spreadsheet(path)

        assert result.profile is not None
        sheet_profile = result.profile.sheet("Profile")
        self.assertEqual(sheet_profile.kind, "form")
        self.assertFalse(sheet_profile.primary_region().header_rows)
        self.assertEqual(
            list(result.sheets["Profile"].columns),
            ["column_1", "column_3"],
        )
        pairs = result.form_summaries["Profile"].to_dict(orient="records")
        self.assertTrue(
            any(row["key"] == "Name" and row["value"] == "Ada" for row in pairs)
        )

    def test_profiles_formulas_chart_and_broken_reference(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "features.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["month", "value", "calculated"])
            sheet.append(["Jan", 10, "=B2*2"])
            sheet.append(["Feb", 20, "=#REF!"])
            chart = BarChart()
            chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3))
            sheet.add_chart(chart, "E2")
            workbook.save(path)

            profile = inspect_workbook(path)

        self.assertEqual(profile.formula_count, 2)
        self.assertEqual(profile.broken_formula_reference_count, 1)
        self.assertEqual(profile.chart_count, 1)

    def test_multiple_merged_header_groups_remain_a_table(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "grouped_report.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            for cell_range, label in (
                ("A1:B1", "Sales"),
                ("C1:D1", "Costs"),
                ("E1:F1", "Meta"),
            ):
                sheet.merge_cells(cell_range)
                sheet[cell_range.split(":")[0]] = label
            sheet.append(["Region", "Revenue", "Labor", "Materials", "Owner", "Year"])
            sheet.append(["East", 10, 2, 3, "Ada", 2026])
            sheet.append(["West", 20, 4, 5, "Lin", 2026])
            workbook.save(path)

            result = read_spreadsheet(path)

        assert result.profile is not None
        sheet_profile = result.profile.sheet("Report")
        self.assertEqual(sheet_profile.kind, "table")
        self.assertEqual(sheet_profile.primary_region().header_rows, (1, 2))

    def test_fragmented_report_preserves_complete_sheet_as_primary_frame(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "fragmented_report.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            for index in range(30):
                row = index * 3 + 1
                sheet.cell(row, 1, f"Item {index}")
                sheet.cell(row, 3, index * 10)
                sheet.cell(row + 1, 1, f"Detail {index}")
                sheet.cell(row + 1, 3, index * 10 + 1)
            workbook.save(path)

            result = read_spreadsheet(path)

        # 原因：格式化报表的空白分隔行会产生许多小区域，最大区域并不代表整张表。
        # 作用：主帧保留全部项目，Agent 可查询任意行；少量区域帧只用于辅助识别局部结构。
        primary = result.sheets["Report"]
        self.assertEqual(result.metadata["Report"]["primary_strategy"], "full_sheet_fallback")
        self.assertTrue(primary.astype(str).eq("Item 29").any().any())
        self.assertTrue(primary.eq(290).any().any())
        self.assertLessEqual(
            len(result.analysis_frames()),
            1 + 4,
        )


if __name__ == "__main__":
    unittest.main()
