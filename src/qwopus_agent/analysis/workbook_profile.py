"""Structural profiling for XLSX workbooks before dataframe analysis."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ConfigDict, Field

MAX_PROFILE_ROWS = 500
MAX_PROFILE_COLUMNS = 160


class TableRegionProfile(BaseModel):
    """One rectangular table candidate detected inside a worksheet."""

    model_config = ConfigDict(frozen=True, extra="forbid")

    region_id: str
    cell_range: str
    min_row: int = Field(ge=1)
    max_row: int = Field(ge=1)
    min_column: int = Field(ge=1)
    max_column: int = Field(ge=1)
    header_rows: tuple[int, ...] = ()
    data_start_row: int | None = None
    column_names: tuple[str, ...] = ()
    confidence: float = Field(default=0.0, ge=0.0, le=1.0)
    non_empty_cells: int = Field(default=0, ge=0)


class SheetProfile(BaseModel):
    """Bounded structure and workbook features for one worksheet."""

    model_config = ConfigDict(frozen=True, extra="forbid")

    name: str
    kind: Literal["empty", "table", "multi_table", "form", "matrix"]
    max_row: int = Field(ge=1)
    max_column: int = Field(ge=1)
    non_empty_cells: int = Field(default=0, ge=0)
    primary_region_id: str | None = None
    table_regions: tuple[TableRegionProfile, ...] = ()
    merged_range_count: int = Field(default=0, ge=0)
    formula_count: int = Field(default=0, ge=0)
    broken_formula_reference_count: int = Field(default=0, ge=0)
    chart_count: int = Field(default=0, ge=0)
    image_count: int = Field(default=0, ge=0)
    data_validation_count: int = Field(default=0, ge=0)
    defined_table_count: int = Field(default=0, ge=0)
    hidden: bool = False
    profile_truncated: bool = False

    def primary_region(self) -> TableRegionProfile | None:
        """Return the explicitly selected primary region."""
        return next(
            (
                region
                for region in self.table_regions
                if region.region_id == self.primary_region_id
            ),
            None,
        )


class WorkbookProfile(BaseModel):
    """Model-neutral workbook structure safe to serialize for tools or UI."""

    model_config = ConfigDict(frozen=True, extra="forbid")

    source: str
    sheet_count: int = Field(ge=0)
    sheets: tuple[SheetProfile, ...] = ()
    formula_count: int = Field(default=0, ge=0)
    broken_formula_reference_count: int = Field(default=0, ge=0)
    merged_range_count: int = Field(default=0, ge=0)
    chart_count: int = Field(default=0, ge=0)
    image_count: int = Field(default=0, ge=0)
    data_validation_count: int = Field(default=0, ge=0)

    def sheet(self, name: str) -> SheetProfile:
        """Resolve one worksheet profile by its exact source name."""
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        raise KeyError(f"Unknown worksheet profile: {name}")


def inspect_workbook(path: Path) -> WorkbookProfile:
    """Inspect an XLSX workbook without evaluating formulas or sending cell data."""
    # 原因：合并区域、公式和绘图对象在 read_only/data_only 模式下不完整。
    # 作用：只在结构检查阶段打开普通工作簿，并关闭 external links 后立即释放。
    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        sheets = tuple(_inspect_sheet(worksheet) for worksheet in workbook.worksheets)
    finally:
        workbook.close()
    return WorkbookProfile(
        source=path.name,
        sheet_count=len(sheets),
        sheets=sheets,
        formula_count=sum(sheet.formula_count for sheet in sheets),
        broken_formula_reference_count=sum(
            sheet.broken_formula_reference_count for sheet in sheets
        ),
        merged_range_count=sum(sheet.merged_range_count for sheet in sheets),
        chart_count=sum(sheet.chart_count for sheet in sheets),
        image_count=sum(sheet.image_count for sheet in sheets),
        data_validation_count=sum(sheet.data_validation_count for sheet in sheets),
    )


def _inspect_sheet(worksheet: Worksheet) -> SheetProfile:
    max_row = max(1, int(worksheet.max_row))
    max_column = max(1, int(worksheet.max_column))
    scan_rows = min(max_row, MAX_PROFILE_ROWS)
    scan_columns = min(max_column, MAX_PROFILE_COLUMNS)
    values = [
        tuple(cell.value for cell in row)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=scan_rows,
            min_col=1,
            max_col=scan_columns,
        )
    ]
    merged_ranges = tuple(worksheet.merged_cells.ranges)
    region_bounds = _detect_region_bounds(values)
    regions = tuple(
        _profile_region(
            worksheet,
            values,
            bounds,
            merged_ranges,
            index=index,
        )
        for index, bounds in enumerate(region_bounds, start=1)
    )
    non_empty_cells = sum(
        1
        for row in values
        for value in row
        if _normalize_cell(value)
    )
    regularity = _row_regularity(values)
    substantial_regions = tuple(
        region
        for region in regions
        if region.non_empty_cells >= 4
        and region.max_row > region.min_row
    )
    kind = _sheet_kind(
        non_empty_cells=non_empty_cells,
        merged_count=len(merged_ranges),
        regularity=regularity,
        regions=substantial_regions,
        form_like=_looks_like_key_value_form(values),
        merged_form_like=_looks_like_merged_form(values, merged_ranges),
    )
    if kind == "form" and regions:
        min_row = min(region.min_row for region in regions)
        max_region_row = max(region.max_row for region in regions)
        min_column = min(region.min_column for region in regions)
        max_region_column = max(region.max_column for region in regions)
        regions = (
            _profile_region(
                worksheet,
                values,
                (min_row, max_region_row, min_column, max_region_column),
                merged_ranges,
                index=1,
            ),
        )
        # 原因：表单的高分文字行不是关系表字段，把它当表头会产生 Unnamed/数值列名。
        # 作用：保留区域边界供原始样本使用，但明确禁止下游按伪表头解释单元格。
        regions = tuple(
            region.model_copy(
                update={
                    "header_rows": (),
                    "data_start_row": region.min_row,
                    "column_names": tuple(
                        f"column_{index}"
                        for index in range(
                            1,
                            region.max_column - region.min_column + 2,
                        )
                    ),
                }
            )
            for region in regions
        )
    selectable_regions = tuple(
        region
        for region in regions
        if region.non_empty_cells >= 4 and region.max_row > region.min_row
    )
    primary = max(
        selectable_regions or regions,
        key=_region_rank,
        default=None,
    )
    formulas = [
        str(value)
        for row in values
        for value in row
        if isinstance(value, str) and value.startswith("=")
    ]
    return SheetProfile(
        name=worksheet.title,
        kind=kind,
        max_row=max_row,
        max_column=max_column,
        non_empty_cells=non_empty_cells,
        primary_region_id=primary.region_id if primary is not None else None,
        table_regions=regions,
        merged_range_count=len(merged_ranges),
        formula_count=len(formulas),
        broken_formula_reference_count=sum("#REF!" in formula for formula in formulas),
        chart_count=len(getattr(worksheet, "_charts", ())),
        image_count=len(getattr(worksheet, "_images", ())),
        data_validation_count=len(worksheet.data_validations.dataValidation),
        defined_table_count=len(worksheet.tables),
        hidden=worksheet.sheet_state != "visible",
        profile_truncated=(
            max_row > MAX_PROFILE_ROWS or max_column > MAX_PROFILE_COLUMNS
        ),
    )


def _detect_region_bounds(
    values: list[tuple[Any, ...]],
) -> tuple[tuple[int, int, int, int], ...]:
    active_rows = [
        row_index
        for row_index, row in enumerate(values, start=1)
        if any(_normalize_cell(value) for value in row)
    ]
    bounds: list[tuple[int, int, int, int]] = []
    for min_row, max_row in _contiguous_ranges(active_rows):
        active_columns = [
            column_index
            for column_index in range(1, len(values[0]) + 1)
            if any(
                _normalize_cell(values[row_index - 1][column_index - 1])
                for row_index in range(min_row, max_row + 1)
            )
        ]
        for min_column, max_column in _contiguous_ranges(active_columns):
            non_empty = sum(
                bool(_normalize_cell(values[row - 1][column - 1]))
                for row in range(min_row, max_row + 1)
                for column in range(min_column, max_column + 1)
            )
            if non_empty >= 2:
                bounds.append((min_row, max_row, min_column, max_column))
    if bounds:
        return tuple(bounds)
    return ()


def _profile_region(
    worksheet: Worksheet,
    values: list[tuple[Any, ...]],
    bounds: tuple[int, int, int, int],
    merged_ranges: tuple[CellRange, ...],
    *,
    index: int,
) -> TableRegionProfile:
    min_row, max_row, min_column, max_column = bounds
    width = max_column - min_column + 1
    candidates: list[tuple[float, int]] = []
    for row_number in range(min_row, max_row):
        current = values[row_number - 1][min_column - 1 : max_column]
        following = values[row_number][min_column - 1 : max_column]
        candidates.append(
            (
                _header_score(
                    current,
                    following,
                    row_number=row_number,
                    min_column=min_column,
                    max_column=max_column,
                    merged_ranges=merged_ranges,
                ),
                row_number,
            )
        )
    best_score = max((candidate[0] for candidate in candidates), default=0.0)
    # 原因：首条数据偶尔缺值时，真实表头会比后续全字符串数据行少一分。
    # 作用：在最高分一分以内优先最早候选；明显低分的标题仍会让位给后续字段行。
    best_row = next(
        (
            row_number
            for score, row_number in candidates
            if score >= best_score - 1.0
        ),
        min_row,
    )
    confidence = min(1.0, max(0.0, best_score / max(1.0, width * 5.0)))
    header_rows: tuple[int, ...] = ()
    if confidence >= 0.45:
        header_rows = (best_row,)
        if _has_group_header(
            worksheet,
            values,
            best_row=best_row,
            min_row=min_row,
            min_column=min_column,
            max_column=max_column,
            merged_ranges=merged_ranges,
        ):
            header_rows = (best_row - 1, best_row)
    column_names = (
        _flatten_column_names(
            worksheet,
            header_rows,
            min_column=min_column,
            max_column=max_column,
            merged_ranges=merged_ranges,
        )
        if header_rows
        else tuple(f"column_{column}" for column in range(1, width + 1))
    )
    non_empty = sum(
        bool(_normalize_cell(values[row - 1][column - 1]))
        for row in range(min_row, max_row + 1)
        for column in range(min_column, max_column + 1)
    )
    cell_range = (
        f"{get_column_letter(min_column)}{min_row}:"
        f"{get_column_letter(max_column)}{max_row}"
    )
    return TableRegionProfile(
        region_id=f"table_{index}",
        cell_range=cell_range,
        min_row=min_row,
        max_row=max_row,
        min_column=min_column,
        max_column=max_column,
        header_rows=header_rows,
        data_start_row=(header_rows[-1] + 1 if header_rows else min_row),
        column_names=column_names,
        confidence=round(confidence, 3),
        non_empty_cells=non_empty,
    )


def _header_score(
    row: tuple[Any, ...],
    next_row: tuple[Any, ...],
    *,
    row_number: int,
    min_column: int,
    max_column: int,
    merged_ranges: tuple[CellRange, ...],
) -> float:
    values = [_normalize_cell(value) for value in row]
    non_empty = [value for value in values if value]
    if (
        len(non_empty) == 1
        and min_column == max_column
        and not _looks_numeric(non_empty[0])
        and any(_looks_numeric(_normalize_cell(value)) for value in next_row)
    ):
        # 原因：合法的单列数值表只有一个字段名，不能满足多列表头的“至少两个文本”规则。
        # 作用：仅在下一行出现数值时识别单列表头，避免普通单列文字清单被误删首项。
        return 5.0
    if len(non_empty) < 2:
        return 0.0
    text_count = sum(not _looks_numeric(value) for value in non_empty)
    unique_count = len(set(non_empty))
    next_count = sum(bool(_normalize_cell(value)) for value in next_row)
    numeric_penalty = len(non_empty) - text_count
    long_penalty = sum(len(value) > 60 for value in non_empty)
    formula_count = sum(value.startswith("=") for value in non_empty)
    merged_title_penalty = 0
    for merged in merged_ranges:
        if (
            merged.min_row == row_number
            and merged.max_row == row_number
            and merged.min_col <= min_column
            and merged.max_col >= max_column
        ):
            merged_title_penalty = 6
            break
    return (
        len(non_empty) * 2
        + text_count * 2
        + unique_count
        + next_count
        - numeric_penalty * 2
        - long_penalty * 2
        # 原因：模板计算区经常连续出现公式，公式字符串并不是字段名称。
        # 作用：降低公式行的表头分数，优先选择上方的人类可读字段行。
        - formula_count * 6
        - merged_title_penalty
    )


def _has_group_header(
    worksheet: Worksheet,
    values: list[tuple[Any, ...]],
    *,
    best_row: int,
    min_row: int,
    min_column: int,
    max_column: int,
    merged_ranges: tuple[CellRange, ...],
) -> bool:
    if best_row <= min_row:
        return False
    previous_row = values[best_row - 2][min_column - 1 : max_column]
    expanded = [
        _merged_cell_value(
            worksheet,
            best_row - 1,
            column,
            merged_ranges,
        )
        for column in range(min_column, max_column + 1)
    ]
    labels = {
        _normalize_cell(value)
        for value in expanded
        if _normalize_cell(value)
    }
    raw_non_empty = sum(bool(_normalize_cell(value)) for value in previous_row)
    # 原因：整行单一合并标题通常是报表名称，不是多层表头。
    # 作用：只有至少两个分组标签的上一行才与字段行合并。
    return raw_non_empty >= 2 and len(labels) >= 2


def _flatten_column_names(
    worksheet: Worksheet,
    header_rows: tuple[int, ...],
    *,
    min_column: int,
    max_column: int,
    merged_ranges: tuple[CellRange, ...],
) -> tuple[str, ...]:
    names: list[str] = []
    for column in range(min_column, max_column + 1):
        parts: list[str] = []
        for row in header_rows:
            label = _normalize_cell(
                _merged_cell_value(worksheet, row, column, merged_ranges)
            )
            if label and (not parts or parts[-1] != label):
                parts.append(label)
        names.append(" | ".join(parts) or f"column_{column - min_column + 1}")
    return _unique_column_names(names)


def _merged_cell_value(
    worksheet: Worksheet,
    row: int,
    column: int,
    merged_ranges: tuple[CellRange, ...],
) -> Any:
    for merged in merged_ranges:
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col <= column <= merged.max_col
        ):
            return worksheet.cell(merged.min_row, merged.min_col).value
    return worksheet.cell(row, column).value


def _sheet_kind(
    *,
    non_empty_cells: int,
    merged_count: int,
    regularity: float,
    regions: tuple[TableRegionProfile, ...],
    form_like: bool,
    merged_form_like: bool,
) -> Literal["empty", "table", "multi_table", "form", "matrix"]:
    if non_empty_cells == 0:
        return "empty"
    if (
        form_like
        or merged_form_like
        or (regularity < 0.35 and len(regions) >= 3)
    ):
        return "form"
    if len(regions) > 1:
        return "multi_table"
    if regions and regions[0].header_rows:
        return "table"
    return "matrix"


def _looks_like_key_value_form(values: list[tuple[Any, ...]]) -> bool:
    non_empty_rows = [
        (index, tuple(value for value in row if _normalize_cell(value)))
        for index, row in enumerate(values)
        if any(_normalize_cell(value) for value in row)
    ]
    if len(non_empty_rows) < 4:
        return False
    title_index, title = non_empty_rows[0]
    if len(title) != 1 or non_empty_rows[1][0] != title_index + 1:
        return False
    candidate_rows = [row for _, row in non_empty_rows[1:] if len(row) == 2]
    if len(candidate_rows) < 3:
        return False
    label_ratio = sum(
        bool(_normalize_cell(row[0])) and not _looks_numeric(_normalize_cell(row[0]))
        for row in candidate_rows
    ) / len(candidate_rows)
    value_text_ratio = sum(
        not _looks_numeric(_normalize_cell(row[1]))
        for row in candidate_rows
    ) / len(candidate_rows)
    return label_ratio >= 0.8 and value_text_ratio >= 0.5


def _looks_like_merged_form(
    values: list[tuple[Any, ...]],
    merged_ranges: tuple[CellRange, ...],
) -> bool:
    matched_rows: set[int] = set()
    for merged in merged_ranges:
        if merged.min_row != merged.max_row or merged.max_col <= merged.min_col:
            continue
        row = merged.min_row
        right_column = merged.max_col + 1
        if row > len(values) or right_column > len(values[row - 1]):
            continue
        label = _normalize_cell(values[row - 1][merged.min_col - 1])
        right_value = _normalize_cell(values[row - 1][right_column - 1])
        if (
            _looks_like_text_label(label)
            and right_value
            and not _cell_is_merged(row, right_column, merged_ranges)
        ):
            matched_rows.add(row)
    # 原因：多组报表表头也常使用合并单元格，不能仅按合并数量判断为表单。
    # 作用：只把至少三行“合并标签 + 普通值单元格”的布局识别为表单。
    return len(matched_rows) >= 3


def _cell_is_merged(
    row: int,
    column: int,
    merged_ranges: tuple[CellRange, ...],
) -> bool:
    return any(
        merged.min_row <= row <= merged.max_row
        and merged.min_col <= column <= merged.max_col
        for merged in merged_ranges
    )


def _looks_like_text_label(value: str) -> bool:
    return bool(value) and not _looks_numeric(value) and len(value) <= 60


def _row_regularity(values: list[tuple[Any, ...]]) -> float:
    counts = [
        sum(bool(_normalize_cell(value)) for value in row)
        for row in values
        if any(_normalize_cell(value) for value in row)
    ]
    if len(counts) < 2:
        return 0.0
    most_common = max(counts.count(count) for count in set(counts))
    return most_common / len(counts)


def _region_rank(region: TableRegionProfile) -> tuple[int, float, int]:
    return (
        region.non_empty_cells,
        region.confidence,
        (region.max_row - region.min_row + 1)
        * (region.max_column - region.min_column + 1),
    )


def _contiguous_ranges(values: list[int]) -> tuple[tuple[int, int], ...]:
    if not values:
        return ()
    ranges: list[tuple[int, int]] = []
    start = previous = values[0]
    for value in values[1:]:
        if value != previous + 1:
            ranges.append((start, previous))
            start = value
        previous = value
    ranges.append((start, previous))
    return tuple(ranges)


def _unique_column_names(names: list[str]) -> tuple[str, ...]:
    counts: dict[str, int] = {}
    unique: list[str] = []
    for name in names:
        count = counts.get(name, 0) + 1
        counts[name] = count
        unique.append(name if count == 1 else f"{name}_{count}")
    return tuple(unique)


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True
