"""Excel-specific loading helpers."""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook as _load_workbook_data_only

from qwopus_agent.analysis.workbook_profile import (
    TableRegionProfile,
    WorkbookProfile,
    inspect_workbook,
)

MAX_ANALYSIS_REGIONS_PER_SHEET = 12
MAX_FRAGMENTED_REGIONS_PER_SHEET = 24
MAX_FRAGMENTED_ANALYSIS_REGIONS_PER_SHEET = 4


@dataclass(frozen=True)
class SpreadsheetReadResult:
    """Loaded spreadsheet sheets plus read metadata."""

    sheets: dict[str, pd.DataFrame]

    metadata: dict[str, dict[str, Any]] = field(default_factory=dict)

    form_summaries: dict[str, pd.DataFrame] = field(default_factory=dict)

    key_value_summaries: dict[str, pd.DataFrame] = field(default_factory=dict)

    region_sheets: dict[str, dict[str, pd.DataFrame]] = field(default_factory=dict)

    profile: WorkbookProfile | None = None

    def analysis_frames(self) -> dict[str, pd.DataFrame]:
        """Return every locally analyzable table under a stable unique name."""
        frames = dict(self.sheets)
        # 原因：form 类型工作表的主 DataFrame 是无结构的泛化列，Agent 无法直接使用。
        # 作用：把已抽取的 key-value 对作为独立帧暴露，让统计和解释能访问真实表单字段。
        for sheet_name, form_summary in self.form_summaries.items():
            if len(form_summary) > 0:
                frames[f"{sheet_name}::form_summary"] = form_summary
        # 原因：角色卡、配置表等工作簿常由多个两列“项目-值”块组成，不适合普通表头解析。
        # 作用：统一暴露为 key_values 帧，支持“年龄是多少”“STR 是多少”这类单项查询。
        for sheet_name, key_value_summary in self.key_value_summaries.items():
            if len(key_value_summary) > 0:
                frames[f"{sheet_name}::key_values"] = key_value_summary
        if self.profile is None:
            return frames
        for sheet_profile in self.profile.sheets:
            sheet_frames = self.region_sheets.get(sheet_profile.name, {})
            region_limit = (
                MAX_FRAGMENTED_ANALYSIS_REGIONS_PER_SHEET
                if len(sheet_profile.table_regions) > MAX_FRAGMENTED_REGIONS_PER_SHEET
                else MAX_ANALYSIS_REGIONS_PER_SHEET
            )
            secondary_regions = sorted(
                (
                    region
                    for region in sheet_profile.table_regions
                    if region.region_id != sheet_profile.primary_region_id
                    and region.non_empty_cells >= 4
                    and len(sheet_frames.get(region.region_id, ())) > 0
                    and len(
                        getattr(sheet_frames.get(region.region_id), "columns", ())
                    )
                    >= 2
                ),
                key=lambda region: (
                    region.non_empty_cells,
                    region.confidence,
                ),
                reverse=True,
            )[:region_limit]
            for region in secondary_regions:
                # 原因：一个工作表可能包含多张独立表，只暴露主区域会让 Agent 永远看不到其他数据。
                # 作用：保留主表名，并以 Sheet::table_N 暴露有数据的高信息次级区域，过滤说明碎片。
                frames[f"{sheet_profile.name}::{region.region_id}"] = sheet_frames[
                    region.region_id
                ]
        return frames


def read_spreadsheet(path: Path) -> SpreadsheetReadResult:
    """Read CSV or Excel with Excel-specific header detection."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return SpreadsheetReadResult(sheets={"csv": pd.read_csv(path)})
    if suffix == ".xlsx":
        return _read_xlsx_with_profiles(path)
    if suffix == ".xls":
        return SpreadsheetReadResult(sheets=pd.read_excel(path, sheet_name=None))
    raise ValueError(f"Unsupported spreadsheet type: {suffix}")


def _read_xlsx_with_profiles(path: Path) -> SpreadsheetReadResult:
    """Read XLSX sheets from an explicit workbook structure profile."""
    profile = inspect_workbook(path)
    raw_sheets = _read_xlsx_data_only(path)
    sheets: dict[str, pd.DataFrame] = {}
    metadata: dict[str, dict[str, Any]] = {}
    form_summaries: dict[str, pd.DataFrame] = {}
    key_value_summaries: dict[str, pd.DataFrame] = {}
    region_sheets: dict[str, dict[str, pd.DataFrame]] = {}

    for sheet_profile in profile.sheets:
        raw = raw_sheets.get(sheet_profile.name, pd.DataFrame())
        form_summary = (
            _extract_form_summary(raw)
            if sheet_profile.kind == "form"
            else pd.DataFrame()
        )
        key_value_summary = (
            _extract_region_key_values(raw, sheet_profile.table_regions)
            if sheet_profile.kind in {"form", "multi_table"}
            else pd.DataFrame()
        )
        frames = {
            region.region_id: _dataframe_from_region(
                raw,
                region,
                use_headers=sheet_profile.kind != "form",
            )
            for region in sheet_profile.table_regions
        }
        primary = sheet_profile.primary_region()
        if (
            len(sheet_profile.table_regions) > MAX_FRAGMENTED_REGIONS_PER_SHEET
            or primary is None
        ):
            dataframe = _generic_dataframe(raw)
        else:
            dataframe = frames[primary.region_id]
            header_region = _prior_header_region(sheet_profile.table_regions, primary)
            if header_region is not None:
                # 原因：带脚注的研究表会把“表头+第一年”和后续年份拆成两个区域。
                # 作用：复用前一区域的字段名并合并数据，避免主帧变成 column_N 且漏掉首段数据。
                dataframe = pd.concat(
                    [
                        frames[header_region.region_id],
                        dataframe.set_axis(list(header_region.column_names), axis=1),
                    ],
                    ignore_index=True,
                )
        sheets[sheet_profile.name] = dataframe
        region_sheets[sheet_profile.name] = frames
        form_summaries[sheet_profile.name] = form_summary
        key_value_summaries[sheet_profile.name] = key_value_summary
        profile_data = sheet_profile.model_dump(mode="json")
        metadata[sheet_profile.name] = {
            "header_row": (
                primary.header_rows[-1]
                if primary is not None and primary.header_rows
                else None
            ),
            "header_rows": list(primary.header_rows) if primary is not None else [],
            "header_detection": "workbook_profile",
            "primary_strategy": (
                "full_sheet_fallback"
                if len(sheet_profile.table_regions) > MAX_FRAGMENTED_REGIONS_PER_SHEET
                else "detected_region"
            ),
            "sheet_kind": sheet_profile.kind,
            "form_pairs": int(len(form_summary)),
            "table_regions": profile_data["table_regions"],
            "primary_region_id": sheet_profile.primary_region_id,
            "formula_count": sheet_profile.formula_count,
            "broken_formula_reference_count": (
                sheet_profile.broken_formula_reference_count
            ),
            "merged_range_count": sheet_profile.merged_range_count,
            "chart_count": sheet_profile.chart_count,
            "image_count": sheet_profile.image_count,
            "data_validation_count": sheet_profile.data_validation_count,
            "profile_truncated": sheet_profile.profile_truncated,
        }

    return SpreadsheetReadResult(
        sheets=sheets,
        metadata=metadata,
        form_summaries=form_summaries,
        key_value_summaries=key_value_summaries,
        region_sheets=region_sheets,
        profile=profile,
    )


def _read_xlsx_data_only(path: Path) -> dict[str, pd.DataFrame]:
    """Read XLSX sheets with computed formula values instead of formula strings."""
    # 原因：pd.read_excel 默认不解析公式缓存值，公式字符串会泄漏到列名和数据中。
    # 作用：用 openpyxl data_only=True 读取 Excel 保存时缓存的计算结果。
    workbook = _load_workbook_data_only(
        path, data_only=True, read_only=True, keep_links=False
    )
    try:
        sheets: dict[str, pd.DataFrame] = {}
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                sheets[worksheet.title] = pd.DataFrame()
                continue
            max_cols = max(len(row) for row in rows)
            padded = [
                tuple(list(row) + [None] * (max_cols - len(row)))
                for row in rows
            ]
            sheets[worksheet.title] = pd.DataFrame(padded)
        return sheets
    finally:
        workbook.close()

def _normalize_cell(value: Any) -> str:
    if value is None or bool(pd.isna(value)):
        return ""
    return str(value).strip()


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def _extract_form_summary(
    dataframe: pd.DataFrame,
    max_rows: int = 200,
    max_columns: int = 80,
    max_pairs: int = 120,
) -> pd.DataFrame:
    """Extract key-value pairs from form-like sheets."""
    rows = [
        tuple(row)
        for row in dataframe.iloc[:max_rows, :max_columns].itertuples(
            index=False,
            name=None,
        )
    ]
    pairs: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()

    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            label = _normalize_cell(value)
            if not _looks_like_label(label):
                continue

            for direction, neighbor in _neighbor_values(rows, row_index, column_index):
                normalized_value = _normalize_cell(neighbor)
                if not normalized_value or normalized_value == label:
                    continue
                identity = (label, normalized_value, direction)
                if identity in seen:
                    continue
                seen.add(identity)
                # 原因：角色卡、申请表这类 Excel 的信息散落在单元格中，不适合强行按二维表头解析。
                # 作用：抽取“标签-值”对，让 LLM 能看到姓名、职业、年龄等表单信息。
                pairs.append(
                    {
                        "key": label,
                        "value": normalized_value,
                        "row": row_index + 1,
                        "column": column_index + 1,
                        "direction": direction,
                    }
                )
                break

            if len(pairs) >= max_pairs:
                return pd.DataFrame(pairs)

    return pd.DataFrame(pairs)


def _extract_region_key_values(
    dataframe: pd.DataFrame,
    regions: Sequence[TableRegionProfile],
    max_pairs: int = 240,
) -> pd.DataFrame:
    """Extract item-value rows from compact two-column regions."""
    pairs: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for region in regions:
        if region.max_column - region.min_column + 1 < 2:
            continue
        window = dataframe.iloc[
            region.min_row - 1 : region.max_row,
            region.min_column - 1 : region.max_column,
        ]
        for row_offset, row in enumerate(window.itertuples(index=False, name=None)):
            cells = [
                (column_offset, _normalize_cell(value), value)
                for column_offset, value in enumerate(row)
                if _normalize_cell(value)
            ]
            if len(cells) < 2:
                continue
            key_column_offset, key, _ = cells[0]
            _, normalized_value, raw_value = cells[1]
            if not _looks_like_label(key) or key == normalized_value:
                continue
            identity = (key, normalized_value, region.region_id)
            if identity in seen:
                continue
            seen.add(identity)
            # 原因：多块角色卡会把首个“键-值”行误判为表头而从普通区域数据中消失。
            # 作用：从原始区域重新建立完整键值索引，保留源表、源行和源列用于追溯。
            pairs.append(
                {
                    "key": key,
                    "value": raw_value,
                    "row": region.min_row + row_offset,
                    "column": region.min_column + key_column_offset,
                    "source_table": region.region_id,
                }
            )
            if len(pairs) >= max_pairs:
                return pd.DataFrame(pairs)
    return pd.DataFrame(pairs)


def _prior_header_region(
    regions: Sequence[TableRegionProfile],
    primary: TableRegionProfile,
) -> TableRegionProfile | None:
    """Find a previous same-width region that can supply headers for the primary body."""
    if primary.header_rows:
        return None
    primary_width = primary.max_column - primary.min_column + 1
    candidates = [
        region
        for region in regions
        if region.max_row < primary.min_row
        and region.min_column == primary.min_column
        and region.max_column == primary.max_column
        and region.header_rows
        and region.max_column - region.min_column + 1 == primary_width
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda region: region.max_row)


def _neighbor_values(
    rows: list[tuple[Any, ...]],
    row_index: int,
    column_index: int,
) -> Iterator[tuple[str, Any]]:
    row = rows[row_index]
    for offset in range(1, 4):
        right_index = column_index + offset
        if right_index >= len(row):
            break
        if _normalize_cell(row[right_index]):
            # 原因：合并标签在 pandas 中会留下一个或多个 NaN 占位列。
            # 作用：跨过这些空列找到同一行的实际表单值，同时限制搜索半径避免串到远处字段。
            direction = "right" if offset == 1 else f"right+{offset}"
            yield direction, row[right_index]
            break
    if row_index + 1 < len(rows):
        next_row = rows[row_index + 1]
        if column_index < len(next_row):
            yield "below", next_row[column_index]


def _looks_like_label(value: str) -> bool:
    if not value or _looks_numeric(value):
        return False
    if len(value) > 40:
        return False
    return any(char.isalpha() or "\u4e00" <= char <= "\u9fff" for char in value)


def _clean_loaded_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    dataframe = dataframe.dropna(axis=0, how="all").dropna(axis=1, how="all")
    dataframe = _drop_example_placeholder_rows(dataframe)
    # 原因：Excel 多行表头会把换行带入真实列名，而 Markdown schema 会把它显示为空格。
    # 作用：让 Agent 从 schema 复制的列名可直接用于沙箱计算，避免同列产生两种表示。
    dataframe.columns = [
        " ".join(str(column).split())
        for column in dataframe.columns
    ]
    # 原因：header=None 会让包含表头文字的原始列先变成 object，切掉表头后不会自动恢复数字类型。
    # 作用：在区域裁剪完成后重新推断 int/float/datetime，保证统计和 pandas 沙箱按真实类型运行。
    return dataframe.infer_objects()


def _drop_example_placeholder_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Remove rows that only mark an example slot and contain no data."""
    if dataframe.empty or dataframe.shape[1] < 2:
        return dataframe
    example_markers = {"eg", "e.g.", "example", "示例", "例"}
    keep_mask: list[bool] = []
    for row in dataframe.itertuples(index=False, name=None):
        first_cell = _normalize_cell(row[0]).casefold()
        remaining_has_data = any(_normalize_cell(value) for value in row[1:])
        # 原因：真实研究表常在表头下放一行 eg，但该行不是观测值。
        # 作用：只删除无数据的示例占位行，避免年份列和统计样本混入文本。
        keep_mask.append(not (first_cell in example_markers and not remaining_has_data))
    return dataframe.loc[keep_mask].copy()


def _dataframe_from_region(
    raw: pd.DataFrame,
    region: TableRegionProfile,
    *,
    use_headers: bool,
) -> pd.DataFrame:
    window = raw.iloc[
        region.min_row - 1 : region.max_row,
        region.min_column - 1 : region.max_column,
    ].copy()
    if use_headers and region.header_rows and region.data_start_row is not None:
        data_offset = max(0, region.data_start_row - region.min_row)
        window = window.iloc[data_offset:].copy()
        # 原因：profiler 用 data_only=False 读取，公式字符串会混入列名。
        # 作用：把以 = 开头的公式列名替换为 column_N，避免 Agent 看到无意义的长公式。
        column_names = [
            name if not str(name).strip().startswith("=")
            else f"column_{i}"
            for i, name in enumerate(region.column_names, start=1)
        ]
        window.columns = column_names
    elif use_headers and len(window) >= 2:
        # 原因：profiler 对低置信度区域返回 headers=()，导致列名退化为 column_N。
        # 作用：当前两行中第一行以文本为主、后续行以数值为主时，推断第一行为表头。
        window = _infer_header_if_present(window)
    else:
        window.columns = list(region.column_names)
    return _clean_loaded_dataframe(window)


def _infer_header_if_present(window: pd.DataFrame) -> pd.DataFrame:
    """Promote the first row to header when it looks like labels over data."""
    if len(window) < 2 or window.shape[1] < 2:
        window.columns = [
            f"column_{i}" for i in range(1, len(window.columns) + 1)
        ]
        return window
    first = window.iloc[0]
    rest = window.iloc[1:]
    text_count = sum(
        bool(_normalize_cell(v)) and not _looks_numeric(_normalize_cell(v))
        for v in first.values
    )
    data_cells = 0
    numeric_cells = 0
    for row in rest.head(5).values:
        for v in row:
            cell = _normalize_cell(v)
            if not cell:
                continue
            data_cells += 1
            if _looks_numeric(cell):
                numeric_cells += 1
    text_ratio = text_count / max(1, window.shape[1])
    numeric_ratio = numeric_cells / max(1, data_cells)
    if text_ratio >= 0.5 and numeric_ratio >= 0.4:
        names = [
            " ".join(str(v).split()) if _normalize_cell(v) else f"column_{i}"
            for i, v in enumerate(first.values, start=1)
        ]
        window = window.iloc[1:].copy()
        window.columns = names
        return window
    window.columns = [
        f"column_{i}" for i in range(1, len(window.columns) + 1)
    ]
    return window


def _generic_dataframe(raw: pd.DataFrame) -> pd.DataFrame:
    dataframe = raw.copy()
    # 原因：碎片化报表需要保留整张表，但第一行仍可能是真实字段名。
    # 作用：在保留全量行的同时尽量恢复可用于统计选择的列名。
    dataframe = _infer_header_if_present(dataframe)
    return _clean_loaded_dataframe(dataframe)
